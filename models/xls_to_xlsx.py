import os
from openpyxl import Workbook, load_workbook

import win32com.client as client

""" Creates the new Workbook """
main_workbook = load_workbook(os.getcwd() + "\\template\\template.xlsx")
main_sheet = main_workbook.active
main_sheet.title = "Summary" # Fill in the date with the last date

""" Converts the .xls files into .xlsm files """
def converter():
    excel = client.Dispatch("excel.application")
    # Converts the files from xls to xlsx
    for file in os.listdir(os.getcwd() + "/old version/"):
        filename, fileextenstion = os.path.splitext(file)
        wb = excel.Workbooks.Open(os.getcwd() + "/old version/" + file)
        output_path = os.getcwd() + "/new version/" + filename
        wb.SaveAs(output_path, 51)
        wb.Close()
    excel.Quit()

""" Searches A column for Headers """
def search_value_in_column(ws, search_string, column='A'):
    for row in range(1, ws.max_row + 1):
        coordinate = "{}{}".format(column, row)
        if ws[coordinate].value == search_string:
            return column, row
    return column, None

""" Searches Row for total """
def search_value_in_row(ws, search_string, row):
    for column in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H','I','K']:
        coordinate = "{}{}".format(column, row)
        if ws[coordinate].value == search_string:
            return column, row
    return row, None

""" Checks which file it currently is scrapping from """
def check_file(files, file_list, coordinate, value):
    if files == file_list[0]:
            main_sheet['B' + coordinate] = value
    elif files == file_list[1]:
        main_sheet['C' + coordinate] = value
    elif files == file_list[2]:
        main_sheet['D' + coordinate] = value
    elif files == file_list[3]:
        main_sheet['E' + coordinate] = value
    elif files == file_list[4]:
        main_sheet['F' + coordinate] = value
    elif files == file_list[5]:
        main_sheet['G' + coordinate] = value
    elif files == file_list[6]:
        main_sheet['H' + coordinate] = value

""" Gets the Date """
def date(files, file_list, ws):
    if files == file_list[0]:
        temp_date = ws['A2'].value
        split = temp_date.split('-')
        date = split[0].strip()
        # Checks where to put date
        main_sheet['A4'] = date

""" Saves the Workbook """
def save():
    # Saves the New Weekly Sales Report
    main_workbook.save("Weekly_Sales.xlsx")
