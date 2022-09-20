import os
from pathlib import Path

import openpyxl
import win32com.client as client
from openpyxl import Workbook, load_workbook

import models.xls_to_xlsx as xls

# Clears the new versions folder 
[f.unlink() for f in Path("C:\\Users\\Chest\\Documents\\GitHub\\Toast-Expense-Weekly_Sales\\new version").glob("*") if f.is_file()]

# Converts the xls file to xlsx format
xls.converter()

# Path for new version folder
dir = os.getcwd() + '/new version/'
# Lists all the files in new version
filess = os.listdir(dir)
# Removes all of the lock files in the new version folder
file_list = [files for files in filess if '~' not in files[0]]
# Loops through all the files
for files in file_list:
    if 'False' in files:
        continue
    if files[0] != '~':
        wb = load_workbook(dir + files)
        # Opens the first sheet in the workbook
        ws = wb['Summary']
        # Gets the Date
        if files == file_list[0]:
            xls.date(files, file_list, ws)

        """ Start of Everything """
        # Locates Payment Summary 
        payment_summary = xls.search_value_in_column(ws, 'Payment Summary')
        # Locates Total in Payment Summary row
        payment_summary_total = xls.search_value_in_row(ws, 'Total', payment_summary[1])
        # Locates Credit in Payment Summary 
        credit_location = xls.search_value_in_column(ws, 'Credit', 'B')
        # Total deposit
        x = credit_location[1]
        y = payment_summary_total[0]
        z = y + str(x)
        total_deposit = ws[z].value
        # Loops through each of the files in order fill in the data
        xls.check_file(files, file_list, '7', total_deposit)

        # Gift Cards SOlD 
        sales_summary = xls.search_value_in_column(ws, 'Sales Summary')
        deferred = xls.search_value_in_row(ws, 'Deferred', sales_summary[1])
        x = int(sales_summary[1]) + 1
        y = deferred[0]
        z = y + str(x)
        if 'INone' != z:
            gift_sold_value = ws[z].value
            gift_sold = float(gift_sold_value.strip('$'))
            # Loops through each of the files in order fill in the data
            if gift_sold > 0:
                xls.check_file(files, file_list, '8',  gift_sold)

    # Gift Cards REDEEMED
    gift_card_coords = xls.search_value_in_column(ws, 'Gift Card', 'B')
    x = gift_card_coords[1]
    if 'None' not in str(x):
        y = payment_summary_total[0]
        z = y + str(x)
        gift_card_total = ws[z].value
        # Loops through each of the files in order fill in the data
        xls.check_file(files, file_list, '9', gift_card_total)

    # CASH SALES
    cash_location = xls.search_value_in_column(ws, 'Cash', 'B')
    x = cash_location[1]
    y = payment_summary_total[0]
    z = y + str(x)
    cash_total = ws[z].value
    # Loops through each of the files in order fill in the data
    if cash_total > 0:
        xls.check_file(files, file_list, '12', cash_total)

    # UBER EATS
    uber_eats = xls.search_value_in_column(ws, 'Uber Eats', 'B')
    x = uber_eats[1]
    if 'None' not in str(x):
        z = 'T' + str(x)
        uber_eats_total = ws[z].value
        # Loops through each of the files in order fill in the data
        if uber_eats_total is None:
            pass
        else:
            if uber_eats_total > 0:
                xls.check_file(files, file_list, '13', uber_eats_total)

    # TIPS PAID
    tips_coords = 'E5'
    gratuity_coords = 'D5'
    tips_value = ws[tips_coords].value
    gratuity_value = ws[gratuity_coords].value
    tips = tips_value[1:]
    gratuity = gratuity_value[1:]
    if ',' in tips:
        tipped = tips.replace(',','')
        if ',' in gratuity:
            gratuitys = gratuity.replace(',','')
            tips_paid = round(float(tipped) + float(gratuitys), 3)
        else:
            tips_paid = round(float(tipped) + float(gratuity),3)
    else:
        tips_paid = round(float(tips) + float(gratuity),3)
    # Loops through each of the files in order fill in the data
    if tips_paid > 0:
        xls.check_file(files, file_list, '14', tips_paid)

    # S-FOOD
    sales_categories = xls.search_value_in_column(ws, 'Sales Categories')
    food = xls.search_value_in_column(ws, 'Food', 'B')
    no_category = xls.search_value_in_column(ws, 'No Category', 'B')
    if 'None' != no_category:
        no_category_coords = 'E' + str(no_category[1])
    else:
        no_category = None
    food_amnt_coords = 'E' + str(food[1])
    na_beverage_amnt_coords = 'E' + str(food[1]- 1)
    if no_category_coords != None:
        no_category_total = ws[no_category_coords].value
    food_total = ws[food_amnt_coords].value
    na_beverage_total = ws[na_beverage_amnt_coords].value
    if no_category_total == None:
        final = food_total + na_beverage_total
    else:
        final = food_total + na_beverage_total + no_category_total
    # Loops through each of the files in order fill in the data
    if final > 0:
        xls.check_file(files, file_list, '15', final)

    #S-BAR
    food_beverage_total = 'E' + str(no_category[1] + 1)
    temp3 = ws[food_beverage_total].value
    sub = temp3 - final
    # Loops through each of the files in order fill in the data
    if sub > 0:
        xls.check_file(files, file_list, '16', sub)


    # COMP PROMO
    comp_item_location = xls.search_value_in_column(ws, 'Comp % Item', 'B')
    early_bird_location = xls.search_value_in_column(ws, 'EARLY BIRD', 'B')
    happy_hr_location = xls.search_value_in_column(ws, 'HAPPY HOUR', 'B')
    open_check_dollar_location = xls.search_value_in_column(ws, 'Open $ Check', 'B')
    comp_percent_check = xls.search_value_in_column(ws, 'Comp % Check', 'B')
    open_percent_check = xls.search_value_in_column(ws, 'Open % Check', 'B')
    checks = 0
    # Finds the location using the row
    if None not in open_percent_check:
        open_percent_checked = 'D' + str(open_percent_check[1])
        checks += ws[open_percent_checked].value
    if None not in comp_percent_check:
        comp_percent_checked = 'D' + str(comp_percent_check[1])
        checks += ws[comp_percent_checked].value
    if None not in comp_item_location:
        comp_item_coords = 'D' + str(comp_item_location[1])
        checks += ws[comp_item_coords].value
    if None not in early_bird_location:
        early_bird_coords = 'D' + str(early_bird_location[1])
        checks += ws[early_bird_coords].value
    if None not in happy_hr_location:
        happy_hr_coords = 'D' + str(happy_hr_location[1])
        checks += ws[happy_hr_coords].value
    if None not in open_check_dollar_location:
        open_check_coords = 'D' + str(open_check_dollar_location[1])
        checks += ws[open_check_coords].value
    # Loops through each of the files in order fill in the data
    xls.check_file(files, file_list, '18', checks)

    # MGR DISCOUNT
    check_discounts = xls.search_value_in_column(ws, 'Check Discounts')
    manager_comp = xls.search_value_in_column(ws, 'Manager Comp - Check', 'B')
    if None not in manager_comp:
        price = 'D' + str(manager_comp[1])
        final = ws[price].value
        # Loops through each of the files in order fill in the data
        xls.check_file(files, file_list, '19', final)

    #SERV.DISCOUNT
    employee_discount = xls.search_value_in_column(ws, 'Employee Discount - Check', 'B')
    if None not in employee_discount:
        price = 'D' + str(employee_discount[1])
        final = ws[price].value
        # Loops through each of the files in order fill in the data
        xls.check_file(files, file_list, '21', final)

    # QSA
    qsa_location = xls.search_value_in_column(ws, 'QSA', 'B')
    if None not in qsa_location:
        qsa_coords = 'D' + str(qsa_location[1])
        qsa = ws[qsa_coords].value
        # Loops through each of the files in order fill in the data
        xls.check_file(files, file_list, '22', qsa)

    # SALES TAX
    price = ws['C5'].value
    takes = price[1:]
    final = float(takes)
    # Loops through each of the files in order fill in the data
    xls.check_file(files, file_list, '23', final)

# Clears the new versions folder
[f.unlink() for f in Path("C:\\Users\\Chest\\Documents\\GitHub\\Toast-Expense-Weekly_Sales\\new version").glob("*") if f.is_file()]
print("I'm All done!")
# Saves the New Weekly Sales Report
xls.save()
